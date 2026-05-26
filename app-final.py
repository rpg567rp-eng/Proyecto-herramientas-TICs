"""
=============================================================
  DC - Sistema de Gestión de Producción
  Versión Final (Completa)
  API REST + IA DeepSeek + Exportación Excel/PDF
=============================================================
"""

from flask import Flask, jsonify, request, send_file, render_template
from flask_cors import CORS
import json
import os
import requests
from datetime import datetime, date
import io

# === EXCEL / PDF ===
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    PDF_OK = True
except ImportError:
    PDF_OK = False

# === CONFIG ===
app = Flask(__name__)
CORS(app)

DATA_FILE = os.path.join(os.path.dirname(__file__), 'data', 'pedidos.json')

# ⚠️ IMPORTANTE: Reemplaza con tu propia API Key de DeepSeek
DEEPSEEK_API_KEY = "AQUI_TU_API_KEY"
DEEPSEEK_URL = "https://api.deepseek.com/chat/completions"

ESTADOS_ORDEN = ["pendiente", "produccion", "calidad", "empaquetado", "despachado"]
LIMITES_RETRASO = {"pendiente": 7, "produccion": 5, "calidad": 2, "empaquetado": 1}


# === HELPERS ===
def cargar_datos():
    with open(DATA_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def guardar_datos(data):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def dias_en_estado(actualizado_en_str):
    try:
        actualizado = datetime.fromisoformat(actualizado_en_str)
        delta = datetime.now() - actualizado
        return delta.days
    except:
        return 0

def es_retrasado(pedido):
    estado = pedido.get('estado', '')
    if estado == 'despachado':
        return False
    limite = LIMITES_RETRASO.get(estado, 999)
    return dias_en_estado(pedido.get('actualizadoEn', '')) > limite

def enriquecer_pedido(pedido):
    pedido['diasEnEstado'] = dias_en_estado(pedido.get('actualizadoEn', ''))
    pedido['retrasado'] = es_retrasado(pedido)
    return pedido

def metricas_actuales(data):
    hoy = date.today().isoformat()
    pedidos = data['pedidos']
    despachados_hoy = [p for p in pedidos if p['estado'] == 'despachado' and p.get('actualizadoEn', '').startswith(hoy)]
    por_finalizar = [p for p in pedidos if p['estado'] != 'despachado']
    retrasados = [p for p in pedidos if es_retrasado(p)]
    return {
        "enviadosHoy": len(despachados_hoy),
        "productosVendidosHoy": sum(p['cantidad'] for p in despachados_hoy),
        "porFinalizar": len(por_finalizar),
        "retrasados": len(retrasados),
        "idsRetrasados": [p['id'] for p in retrasados]
    }


# === RUTAS PRINCIPALES ===

@app.route('/')
def index():
    return render_template('index.html')


# === PEDIDOS ===

@app.route('/api/pedidos', methods=['GET'])
def get_pedidos():
    data = cargar_datos()
    estado_filtro = request.args.get('estado', None)
    pedidos = data['pedidos']
    if estado_filtro and estado_filtro != 'todos':
        pedidos = [p for p in pedidos if p['estado'] == estado_filtro]
    pedidos = [enriquecer_pedido(p) for p in pedidos]
    pedidos.sort(key=lambda p: p.get('fechaEntrega', ''))
    return jsonify({"pedidos": pedidos, "total": len(pedidos)})


@app.route('/api/pedidos', methods=['POST'])
def crear_pedido():
    data = cargar_datos()
    body = request.get_json()

    requeridos = ['cliente', 'producto', 'cantidad', 'fechaEntrega']
    for campo in requeridos:
        if not body.get(campo):
            return jsonify({"error": f"El campo '{campo}' es requerido"}), 400

    producto_nombre = body['producto']
    cantidad = int(body['cantidad'])
    inventario = data['inventario']
    prod = next((p for p in inventario if p['nombre'] == producto_nombre), None)

    if prod is None:
        return jsonify({"error": "Producto no encontrado"}), 404
    if prod['stock'] < cantidad:
        return jsonify({"error": f"Stock insuficiente: solo quedan {prod['stock']} unidades"}), 400

    prod['stock'] -= cantidad

    data['contador'] += 1
    nuevo_id = f"DC-{data['contador']:04d}"
    ahora = datetime.now().isoformat()

    nuevo_pedido = {
        "id": nuevo_id,
        "cliente": body['cliente'],
        "telefono": body.get('telefono', ''),
        "producto": producto_nombre,
        "cantidad": cantidad,
        "fechaEntrega": body['fechaEntrega'],
        "estado": body.get('estado', 'pendiente'),
        "notas": body.get('notas', ''),
        "creadoEn": ahora,
        "actualizadoEn": ahora
    }

    data['pedidos'].append(nuevo_pedido)
    guardar_datos(data)

    return jsonify({"success": True, "pedido": enriquecer_pedido(nuevo_pedido), "id": nuevo_id}), 201


@app.route('/api/pedidos/<pedido_id>', methods=['PUT'])
def actualizar_pedido(pedido_id):
    data = cargar_datos()
    body = request.get_json()
    pedido = next((p for p in data['pedidos'] if p['id'] == pedido_id), None)

    if not pedido:
        return jsonify({"error": "Pedido no encontrado"}), 404

    campos = ['estado', 'notas', 'fechaEntrega', 'cliente', 'telefono']
    for campo in campos:
        if campo in body:
            pedido[campo] = body[campo]

    pedido['actualizadoEn'] = datetime.now().isoformat()
    guardar_datos(data)
    return jsonify({"success": True, "pedido": enriquecer_pedido(pedido)})


@app.route('/api/pedidos/<pedido_id>/avanzar', methods=['POST'])
def avanzar_estado(pedido_id):
    data = cargar_datos()
    pedido = next((p for p in data['pedidos'] if p['id'] == pedido_id), None)

    if not pedido:
        return jsonify({"error": "Pedido no encontrado"}), 404

    estado_actual = pedido['estado']
    if estado_actual == 'despachado':
        return jsonify({"error": "El pedido ya fue despachado"}), 400

    idx = ESTADOS_ORDEN.index(estado_actual)
    pedido['estado'] = ESTADOS_ORDEN[idx + 1]
    pedido['actualizadoEn'] = datetime.now().isoformat()
    guardar_datos(data)
    return jsonify({"success": True, "pedido": enriquecer_pedido(pedido), "nuevoEstado": pedido['estado']})


@app.route('/api/pedidos/<pedido_id>', methods=['DELETE'])
def eliminar_pedido(pedido_id):
    data = cargar_datos()
    pedido = next((p for p in data['pedidos'] if p['id'] == pedido_id), None)

    if not pedido:
        return jsonify({"error": "Pedido no encontrado"}), 404

    if pedido['estado'] != 'despachado':
        inv = next((p for p in data['inventario'] if p['nombre'] == pedido['producto']), None)
        if inv:
            inv['stock'] += pedido['cantidad']

    data['pedidos'] = [p for p in data['pedidos'] if p['id'] != pedido_id]
    guardar_datos(data)
    return jsonify({"success": True})


# === MÉTRICAS ===

@app.route('/api/metricas', methods=['GET'])
def get_metricas():
    data = cargar_datos()
    return jsonify(metricas_actuales(data))


# === INVENTARIO ===

@app.route('/api/inventario', methods=['GET'])
def get_inventario():
    data = cargar_datos()
    return jsonify({"inventario": data['inventario']})


@app.route('/api/inventario/<int:prod_id>', methods=['PUT'])
def actualizar_stock(prod_id):
    data = cargar_datos()
    body = request.get_json()
    prod = next((p for p in data['inventario'] if p['id'] == prod_id), None)
    if not prod:
        return jsonify({"error": "Producto no encontrado"}), 404
    if 'stock' in body:
        prod['stock'] = int(body['stock'])
    guardar_datos(data)
    return jsonify({"success": True, "producto": prod})


# === CHATBOT IA ===

@app.route('/api/chat', methods=['POST'])
def chat():
    body = request.get_json()
    mensaje_usuario = body.get('mensaje', '')
    historial = body.get('historial', [])

    data = cargar_datos()
    metricas = metricas_actuales(data)
    pedidos = data['pedidos']
    inventario = data['inventario']

    pedidos_activos = [p for p in pedidos if p['estado'] != 'despachado']
    retrasados = [p for p in pedidos if es_retrasado(p)]

    contexto_sistema = f"""Eres el asistente de producción de DC, una empresa colombiana de regalos creativos artesanales.

DATOS EN TIEMPO REAL (HOY {date.today().strftime('%d/%m/%Y')}):
- Pedidos enviados hoy: {metricas['enviadosHoy']}
- Productos vendidos hoy: {metricas['productosVendidosHoy']}
- Pedidos por finalizar: {metricas['porFinalizar']}
- Pedidos retrasados: {metricas['retrasados']} ({', '.join(metricas['idsRetrasados']) if metricas['idsRetrasados'] else 'ninguno'})

PEDIDOS ACTIVOS ({len(pedidos_activos)}):
{chr(10).join([f"- {p['id']}: {p['producto']} x{p['cantidad']} para {p['cliente']} — Estado: {p['estado']} — Entrega: {p['fechaEntrega']}" for p in pedidos_activos[:10]])}

INVENTARIO:
{chr(10).join([f"- {p['emoji']} {p['nombre']}: {p['stock']} unidades {'⚠️ STOCK BAJO' if p['stock'] <= p['stockMinimo'] else ''}" for p in inventario])}

INSTRUCCIONES:
- Responde SIEMPRE en español, de forma concisa y útil.
- Usa emojis relevantes.
- NO inventes datos: usa solo la información del contexto."""

    mensajes = [{"role": "system", "content": contexto_sistema}]
    for h in historial[-10:]:
        mensajes.append({"role": h['rol'], "content": h['contenido']})
    mensajes.append({"role": "user", "content": mensaje_usuario})

    try:
        resp = requests.post(
            DEEPSEEK_URL,
            headers={
                "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
                "Content-Type": "application/json"
            },
            json={
                "model": "deepseek-chat",
                "messages": mensajes,
                "max_tokens": 400,
                "temperature": 0.7
            },
            timeout=15
        )
        resp.raise_for_status()
        resultado = resp.json()
        respuesta_ia = resultado['choices'][0]['message']['content']
        return jsonify({"respuesta": respuesta_ia, "ok": True})

    except requests.exceptions.Timeout:
        return jsonify({"respuesta": "⏱️ La IA tardó demasiado. Intenta de nuevo.", "ok": False})
    except Exception as e:
        return jsonify({"respuesta": f"⚠️ Error: {str(e)[:80]}", "ok": False})


# === EXPORTAR EXCEL ===

@app.route('/api/exportar/excel', methods=['GET'])
def exportar_excel():
    if not EXCEL_OK:
        return jsonify({"error": "pip install openpyxl"}), 500

    data = cargar_datos()
    pedidos = data['pedidos']
    metricas = metricas_actuales(data)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Pedidos"

    header_fill = PatternFill("solid", fgColor="5A7A2B")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin', color='E0DDD3'),
        right=Side(style='thin', color='E0DDD3'),
        top=Side(style='thin', color='E0DDD3'),
        bottom=Side(style='thin', color='E0DDD3')
    )

    encabezados = ["# Pedido", "Cliente", "Teléfono", "Producto", "Cantidad", "Fecha Entrega", "Estado", "Días", "Retrasado", "Notas"]
    ws1.append(encabezados)

    for col_num, _ in enumerate(encabezados, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for pedido in pedidos:
        dias = dias_en_estado(pedido.get('actualizadoEn', ''))
        retrasado = "SÍ" if es_retrasado(pedido) else "No"
        fila = [
            pedido['id'], pedido['cliente'], pedido.get('telefono',''),
            pedido['producto'], pedido['cantidad'], pedido['fechaEntrega'],
            pedido['estado'].upper(), dias, retrasado, pedido.get('notas','')
        ]
        ws1.append(fila)

    ws2 = wb.create_sheet("Inventario")
    ws2.append(["Producto", "Emoji", "Stock", "Stock Mínimo", "Stock Inicial", "Estado"])
    for col in range(1, 7):
        ws2.cell(row=1, column=col).fill = header_fill
        ws2.cell(row=1, column=col).font = header_font

    for prod in data['inventario']:
        estado = "BAJO" if prod['stock'] <= prod['stockMinimo'] else "OK"
        ws2.append([prod['nombre'], prod['emoji'], prod['stock'], prod['stockMinimo'], prod['stockInicial'], estado])

    ws3 = wb.create_sheet("Resumen")
    ws3['A1'] = "DC — Resumen"
    ws3['A1'].font = Font(bold=True, size=14)
    ws3.append(["Métrica", "Valor"])
    ws3.append(["Enviados hoy", metricas['enviadosHoy']])
    ws3.append(["Productos vendidos", metricas['productosVendidosHoy']])
    ws3.append(["Por finalizar", metricas['porFinalizar']])
    ws3.append(["Retrasados", metricas['retrasados']])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"DC_Pedidos_{date.today().isoformat()}.xlsx")


# === EXPORTAR PDF ===

@app.route('/api/exportar/pdf', methods=['GET'])
def exportar_pdf():
    if not PDF_OK:
        return jsonify({"error": "pip install reportlab"}), 500

    data = cargar_datos()
    pedidos = data['pedidos']
    metricas = metricas_actuales(data)

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    verde = colors.HexColor('#5a7a2b')

    story = []
    story.append(Paragraph("DC — Reporte de Pedidos", ParagraphStyle('titulo', fontSize=20, textColor=verde, fontName='Helvetica-Bold')))
    story.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 0.5*cm))

    # Métricas
    metricas_data = [["Enviados", "Vendidos", "Pendientes", "Retrasados"],
                     [str(metricas['enviadosHoy']), str(metricas['productosVendidosHoy']),
                      str(metricas['porFinalizar']), str(metricas['retrasados'])]]
    t_metricas = Table(metricas_data, colWidths=[4*cm, 4*cm, 4*cm, 4*cm])
    t_metricas.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), verde),
                                    ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e0ddd3'))]))
    story.append(t_metricas)
    story.append(Spacer(1, 0.5*cm))

    # Tabla de pedidos
    encabezados_pdf = ["ID", "Cliente", "Producto", "Cant.", "Entrega", "Estado"]
    filas_pdf = [encabezados_pdf]
    for p in pedidos[:20]:
        filas_pdf.append([p['id'], p['cliente'][:15], p['producto'][:15], str(p['cantidad']), p['fechaEntrega'], p['estado']])

    t_pedidos = Table(filas_pdf, colWidths=[2*cm, 3.5*cm, 4*cm, 1.5*cm, 2.5*cm, 2.5*cm])
    t_pedidos.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), verde),
                                   ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                                   ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                                   ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#e0ddd3'))]))
    story.append(t_pedidos)

    doc.build(story)
    output.seek(0)

    return send_file(output, mimetype='application/pdf', as_attachment=True, download_name=f"DC_Reporte_{date.today().isoformat()}.pdf")


# === INICIAR SERVIDOR ===
if __name__ == '__main__':
    print("\n🍬 DC — Servidor iniciando...")
    print("   http://localhost:5000")
    print("   Ctrl+C para detener\n")
    app.run(debug=True, port=5000)